from rest_framework import status
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from apps.core.api import TenantRequiredMixin
from apps.tenancy.models import Membership
from apps.tenancy.permissions import IsTenantMember, HasTenantRole
from apps.core.audit_mixins import AuditCrudMixin
from apps.core.audit import log_audit_event
from .models import Warehouse, Item, StockEntry, StockExit
from .models import StockLayer, InventoryAdjustment
from .models import StockTransfer, UFVRate
from .models import UFVRevaluationRun, UFVRevaluationRunLine
from .models import (
    StockEntryDocument,
    StockEntryLine,
    StockExitDocument,
    StockExitLine,
    StockExitLineAllocation,
    DocumentType,
)
from .serializers import StockTransferSerializer
from .serializers import WarehouseSerializer
from .serializers import ItemSerializer
from .serializers import StockEntrySerializer
from .serializers import StockExitSerializer
from .serializers import InventoryAdjustmentSerializer
from .serializers import UFVRateSerializer
from .serializers import UFVRevaluationRunSerializer
from .serializers import (
    StockEntryDocumentSerializer,
    StockExitDocumentSerializer,
    DocumentTypeSerializer,
)
from django.db import transaction, IntegrityError
from django.db.models import Sum, F, DecimalField, ExpressionWrapper
from decimal import Decimal
from django.utils import timezone
from django.http import FileResponse
from rest_framework.decorators import action

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.core.excel import add_tenant_report_header
from apps.core.pdf import build_inventory_valuation_pdf
from apps.core.pdf import build_current_stock_pdf
from apps.core.pdf import build_kardex_pdf
from apps.core.pdf import build_ufv_revaluation_run_pdf

from apps.inventory.document_pdfs import (
    build_stock_entry_document_pdf,
    build_stock_exit_document_pdf,
)


class WarehouseViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = WarehouseSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]
    required_roles = [Membership.Role.OWNER, Membership.Role.ADMIN]

    def get_queryset(self):
        # Se evalúa ya dentro del request, con tenant en el contexto ✅
        return Warehouse.objects.all()


class ItemViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]
    required_roles = [Membership.Role.OWNER, Membership.Role.ADMIN]

    def get_queryset(self):
        return Item.objects.all()


class StockEntryViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = StockEntrySerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]
    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
        Membership.Role.MEMBER,
    ]

    def get_queryset(self):
        return StockEntry.objects.select_related("warehouse", "item").all()
    
    def perform_create(self, serializer):
        tenant = self.request.tenant
        entry_date = serializer.validated_data.get("entry_date")

        ufv_rate = (
            UFVRate.objects
            .filter(tenant=tenant, date=entry_date)
            .first()
        )

        stock_entry = serializer.save(
            tenant=tenant,
            ufv_rate=ufv_rate,
            ufv_value=ufv_rate.value if ufv_rate else None,
        )

        StockLayer.objects.create(
            tenant=tenant,
            stock_entry=stock_entry,
            warehouse=stock_entry.warehouse,
            item=stock_entry.item,
            original_quantity=stock_entry.quantity,
            remaining_quantity=stock_entry.quantity,
            unit_cost=stock_entry.unit_cost,
            entry_date=stock_entry.entry_date,
            ufv_value=stock_entry.ufv_value,
        )


class StockExitViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = StockExitSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]
    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
        Membership.Role.MEMBER,
    ]

    def get_queryset(self):
        return (
            StockExit.objects
            .select_related("warehouse", "item")
            .prefetch_related("allocations")
            .all()
        )


class CurrentStockView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get(self, request):
        layer_total_cost = ExpressionWrapper(
            F("remaining_quantity") * F("unit_cost"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )

        rows = (
            StockLayer.objects
            .select_related("warehouse", "item")
            .values(
                "warehouse_id",
                "warehouse__name",
                "item_id",
                "item__code",
                "item__name",
            )
            .annotate(
                quantity=Sum("remaining_quantity"),
                total_cost=Sum(layer_total_cost),
            )
        )

        result = []

        for row in rows:
            quantity = row["quantity"] or Decimal("0")
            total_cost = row["total_cost"] or Decimal("0")

            average_cost = (
                total_cost / quantity
                if quantity > 0
                else Decimal("0")
            )

            result.append({
                "warehouse_id": row["warehouse_id"],
                "warehouse_name": row["warehouse__name"],
                "item_id": row["item_id"],
                "item_code": row["item__code"],
                "item_name": row["item__name"],
                "quantity": str(quantity.quantize(Decimal("0.01"))),
                "average_cost": str(average_cost.quantize(Decimal("0.01"))),
                "total_cost": str(total_cost.quantize(Decimal("0.01"))),
            })

        return Response(result)


class DashboardSummaryView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get(self, request):
        total_warehouses = Warehouse.objects.count()
        total_items = Item.objects.count()
        total_entries = StockEntry.objects.count()
        total_exits = StockExit.objects.count()

        entries_quantity = (
            StockEntry.objects.aggregate(total=Sum("quantity")).get("total")
            or Decimal("0")
        )

        exits_quantity = (
            StockExit.objects.aggregate(total=Sum("quantity")).get("total")
            or Decimal("0")
        )

        current_quantity = entries_quantity - exits_quantity

        entry_total_cost = ExpressionWrapper(
            F("quantity") * F("unit_cost"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )

        total_entry_cost = (
            StockEntry.objects
            .aggregate(total=Sum(entry_total_cost))
            .get("total")
            or Decimal("0")
        )

        layer_total_cost = ExpressionWrapper(
            F("remaining_quantity") * F("unit_cost"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )

        current_quantity = (
            StockLayer.objects.aggregate(total=Sum("remaining_quantity")).get("total")
            or Decimal("0")
        )

        current_value = (
            StockLayer.objects.aggregate(total=Sum(layer_total_cost)).get("total")
            or Decimal("0")
        )

        return Response({
            "total_warehouses": total_warehouses,
            "total_items": total_items,
            "total_stock_entries": total_entries,
            "total_stock_exits": total_exits,
            "current_quantity": str(current_quantity),
            "current_value": str(total_entry_cost),
            "current_quantity": str(current_quantity.quantize(Decimal("0.01"))),
            "current_value": str(current_value.quantize(Decimal("0.01"))),
        })


class StockMovementHistoryView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get(self, request):
        entries = [
            {
                "type": "ENTRY",
                "date": entry.entry_date,
                "warehouse_name": entry.warehouse.name,
                "item_code": entry.item.code,
                "item_name": entry.item.name,
                "quantity": str(entry.quantity),
                "unit_cost": str(entry.unit_cost),
                "total_cost": str((entry.quantity * entry.unit_cost).quantize(Decimal("0.01"))),
                "reference": entry.reference,
                "notes": entry.notes,
            }
            for entry in StockEntry.objects.select_related("warehouse", "item").all()
        ]

        exits = []

        for stock_exit in (
            StockExit.objects
            .select_related("warehouse", "item")
            .prefetch_related("allocations")
            .all()
        ):
            total_cost = sum(
                (
                    allocation.quantity * allocation.unit_cost
                    for allocation in stock_exit.allocations.all()
                ),
                Decimal("0"),
            )

            exits.append({
                "type": "EXIT",
                "date": stock_exit.exit_date,
                "warehouse_name": stock_exit.warehouse.name,
                "item_code": stock_exit.item.code,
                "item_name": stock_exit.item.name,
                "quantity": str(stock_exit.quantity),
                "unit_cost": None,
                "total_cost": str(total_cost.quantize(Decimal("0.01"))),
                "reference": stock_exit.reference,
                "notes": stock_exit.notes,
            })

        adjustments = []

        for adjustment in (
            InventoryAdjustment.objects
            .select_related("warehouse", "item")
            .prefetch_related("allocations")
            .all()
        ):
            if adjustment.adjustment_type == InventoryAdjustment.TYPE_POSITIVE:
                total_cost = adjustment.quantity * adjustment.unit_cost
                unit_cost = adjustment.unit_cost
            else:
                total_cost = sum(
                    (
                        allocation.quantity * allocation.unit_cost
                        for allocation in adjustment.allocations.all()
                    ),
                    Decimal("0"),
                )
                unit_cost = None

            adjustments.append({
                "type": f"ADJUSTMENT_{adjustment.adjustment_type}",
                "date": adjustment.adjustment_date,
                "warehouse_name": adjustment.warehouse.name,
                "item_code": adjustment.item.code,
                "item_name": adjustment.item.name,
                "quantity": str(adjustment.quantity),
                "unit_cost": str(unit_cost) if unit_cost is not None else None,
                "total_cost": str(total_cost.quantize(Decimal("0.01"))),
                "reference": adjustment.reference,
                "notes": adjustment.reason,
            })
        
        transfers = []

        for transfer in (
            StockTransfer.objects
            .select_related("source_warehouse", "destination_warehouse", "item")
            .prefetch_related("allocations")
            .all()
        ):
            total_cost = sum(
                (
                    allocation.quantity * allocation.unit_cost
                    for allocation in transfer.allocations.all()
                ),
                Decimal("0"),
            )

            transfers.append({
                "type": "TRANSFER",
                "date": transfer.transfer_date,
                "warehouse_name": f"{transfer.source_warehouse.name} → {transfer.destination_warehouse.name}",
                "item_code": transfer.item.code,
                "item_name": transfer.item.name,
                "quantity": str(transfer.quantity),
                "unit_cost": None,
                "total_cost": str(total_cost.quantize(Decimal("0.01"))),
                "reference": transfer.reference,
                "notes": transfer.notes,
            })

        entry_documents = []

        for document in (
            StockEntryDocument.objects
            .prefetch_related("lines__warehouse", "lines__item")
            .filter(status=StockEntryDocument.Status.CONFIRMED)
            .order_by("-entry_date", "-id")[:20]
        ):
            for line in document.lines.all():
                entry_documents.append({
                    "type": "ENTRY_DOCUMENT",
                    "date": document.entry_date,
                    "warehouse_name": line.warehouse.name,
                    "item_code": line.item.code,
                    "item_name": line.item.name,
                    "quantity": str(line.quantity),
                    "unit_cost": str(line.unit_cost),
                    "total_cost": str(line.total_cost.quantize(Decimal("0.01"))),
                    "reference": document.document_number,
                    "notes": document.reason,
                })
        
        exit_documents = []

        for document in (
            StockExitDocument.objects
            .prefetch_related("lines__warehouse", "lines__item")
            .filter(status=StockExitDocument.Status.CONFIRMED)
            .order_by("-exit_date", "-id")[:20]
        ):
            for line in document.lines.all():
                exit_documents.append({
                    "type": "EXIT_DOCUMENT",
                    "date": document.exit_date,
                    "warehouse_name": line.warehouse.name,
                    "item_code": line.item.code,
                    "item_name": line.item.name,
                    "quantity": str(line.quantity),
                    "unit_cost": None,
                    "total_cost": str(line.total_cost.quantize(Decimal("0.01"))),
                    "reference": document.document_number,
                    "notes": document.reason,
                })

        # movements = entries + exits + adjustments + transfers
        movements = (
            entries
            + exits
            + adjustments
            + transfers
            + entry_documents
            + exit_documents
        )
        movements = sorted(
            movements,
            key=lambda movement: movement["date"],
            reverse=True,
        )[:10]

        return Response(movements)


class KardexView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get(self, request):
        warehouse_id = request.query_params.get("warehouse")
        item_id = request.query_params.get("item")

        if not warehouse_id or not item_id:
            return Response(
                {"detail": "warehouse and item query params are required."},
                status=400,
            )

        movements = []

        entries = (
            StockEntry.objects
            .select_related("warehouse", "item")
            .filter(warehouse_id=warehouse_id, item_id=item_id)
        )

        for entry in entries:
            movements.append({
                "date": entry.entry_date,
                "type": "ENTRY",
                "reference": entry.reference,
                "entry_quantity": entry.quantity,
                "exit_quantity": Decimal("0"),
                "unit_cost": entry.unit_cost,
                "total_cost": entry.quantity * entry.unit_cost,
                "sort_id": entry.id,
            })
        
        entry_documents = (
            StockEntryDocument.objects
            .prefetch_related("lines__warehouse", "lines__item")
            .filter(
                status=StockEntryDocument.Status.CONFIRMED,
                lines__warehouse_id=warehouse_id,
                lines__item_id=item_id,
            )
            .distinct()
        )

        for document in entry_documents:
            for line in document.lines.all():
                if str(line.warehouse_id) != str(warehouse_id) or str(line.item_id) != str(item_id):
                    continue

                movements.append({
                    "date": document.entry_date,
                    "type": "ENTRY_DOCUMENT",
                    "reference": document.document_number,
                    "entry_quantity": line.quantity,
                    "exit_quantity": Decimal("0"),
                    "unit_cost": line.unit_cost,
                    "total_cost": line.total_cost,
                    "sort_id": line.id,
                })

        exits = (
            StockExit.objects
            .select_related("warehouse", "item")
            .prefetch_related("allocations")
            .filter(warehouse_id=warehouse_id, item_id=item_id)
        )

        for stock_exit in exits:
            total_cost = sum(
                (
                    allocation.quantity * allocation.unit_cost
                    for allocation in stock_exit.allocations.all()
                ),
                Decimal("0"),
            )

            average_exit_cost = (
                total_cost / stock_exit.quantity
                if stock_exit.quantity > 0
                else Decimal("0")
            )

            movements.append({
                "date": stock_exit.exit_date,
                "type": "EXIT",
                "reference": stock_exit.reference,
                "entry_quantity": Decimal("0"),
                "exit_quantity": stock_exit.quantity,
                "unit_cost": average_exit_cost,
                "total_cost": total_cost,
                "sort_id": stock_exit.id,
            })
        
        exit_documents = (
            StockExitDocument.objects
            .prefetch_related("lines__warehouse", "lines__item", "lines__allocations")
            .filter(
                status=StockExitDocument.Status.CONFIRMED,
                lines__warehouse_id=warehouse_id,
                lines__item_id=item_id,
            )
            .distinct()
        )

        for document in exit_documents:
            for line in document.lines.all():
                if str(line.warehouse_id) != str(warehouse_id) or str(line.item_id) != str(item_id):
                    continue

                movements.append({
                    "date": document.exit_date,
                    "type": "EXIT_DOCUMENT",
                    "reference": document.document_number,
                    "entry_quantity": Decimal("0"),
                    "exit_quantity": line.quantity,
                    "unit_cost": (
                        line.total_cost / line.quantity
                        if line.quantity > 0
                        else Decimal("0")
                    ),
                    "total_cost": line.total_cost,
                    "sort_id": line.id,
                })
        
        adjustments = (
            InventoryAdjustment.objects
            .select_related("warehouse", "item")
            .prefetch_related("allocations")
            .filter(warehouse_id=warehouse_id, item_id=item_id)
        )

        for adjustment in adjustments:
            if adjustment.adjustment_type == InventoryAdjustment.TYPE_POSITIVE:
                total_cost = adjustment.quantity * adjustment.unit_cost

                movements.append({
                    "date": adjustment.adjustment_date,
                    "type": "ADJUSTMENT_POSITIVE",
                    "reference": adjustment.reference,
                    "entry_quantity": adjustment.quantity,
                    "exit_quantity": Decimal("0"),
                    "unit_cost": adjustment.unit_cost,
                    "total_cost": total_cost,
                    "sort_id": adjustment.id,
                })

            if adjustment.adjustment_type == InventoryAdjustment.TYPE_NEGATIVE:
                total_cost = sum(
                    (
                        allocation.quantity * allocation.unit_cost
                        for allocation in adjustment.allocations.all()
                    ),
                    Decimal("0"),
                )

                average_exit_cost = (
                    total_cost / adjustment.quantity
                    if adjustment.quantity > 0
                    else Decimal("0")
                )

                movements.append({
                    "date": adjustment.adjustment_date,
                    "type": "ADJUSTMENT_NEGATIVE",
                    "reference": adjustment.reference,
                    "entry_quantity": Decimal("0"),
                    "exit_quantity": adjustment.quantity,
                    "unit_cost": average_exit_cost,
                    "total_cost": total_cost,
                    "sort_id": adjustment.id,
                })
        
        transfers_out = (
            StockTransfer.objects
            .select_related("source_warehouse", "destination_warehouse", "item")
            .prefetch_related("allocations")
            .filter(source_warehouse_id=warehouse_id, item_id=item_id)
        )

        for transfer in transfers_out:
            total_cost = sum(
                (
                    allocation.quantity * allocation.unit_cost
                    for allocation in transfer.allocations.all()
                ),
                Decimal("0"),
            )

            average_cost = (
                total_cost / transfer.quantity
                if transfer.quantity > 0
                else Decimal("0")
            )

            movements.append({
                "date": transfer.transfer_date,
                "type": "TRANSFER_OUT",
                "reference": transfer.reference,
                "entry_quantity": Decimal("0"),
                "exit_quantity": transfer.quantity,
                "unit_cost": average_cost,
                "total_cost": total_cost,
                "sort_id": transfer.id,
            })

        transfers_in = (
            StockTransfer.objects
            .select_related("source_warehouse", "destination_warehouse", "item")
            .prefetch_related("allocations")
            .filter(destination_warehouse_id=warehouse_id, item_id=item_id)
        )

        for transfer in transfers_in:
            total_cost = sum(
                (
                    allocation.quantity * allocation.unit_cost
                    for allocation in transfer.allocations.all()
                ),
                Decimal("0"),
            )

            average_cost = (
                total_cost / transfer.quantity
                if transfer.quantity > 0
                else Decimal("0")
            )

            movements.append({
                "date": transfer.transfer_date,
                "type": "TRANSFER_IN",
                "reference": transfer.reference,
                "entry_quantity": transfer.quantity,
                "exit_quantity": Decimal("0"),
                "unit_cost": average_cost,
                "total_cost": total_cost,
                "sort_id": transfer.id,
            })

        movements.sort(key=lambda row: (row["date"], row["sort_id"]))

        balance_quantity = Decimal("0")
        balance_value = Decimal("0")
        result = []

        for row in movements:
            if row["type"] == "ENTRY":
                balance_quantity += row["entry_quantity"]
                balance_value += row["total_cost"]

            if row["type"] == "EXIT":
                balance_quantity -= row["exit_quantity"]
                balance_value -= row["total_cost"]

            average_balance_cost = (
                balance_value / balance_quantity
                if balance_quantity > 0
                else Decimal("0")
            )

            result.append({
                "date": row["date"],
                "type": row["type"],
                "reference": row["reference"],
                "entry_quantity": str(row["entry_quantity"].quantize(Decimal("0.01"))),
                "exit_quantity": str(row["exit_quantity"].quantize(Decimal("0.01"))),
                "balance_quantity": str(balance_quantity.quantize(Decimal("0.01"))),
                "unit_cost": str(row["unit_cost"].quantize(Decimal("0.01"))),
                "total_cost": str(row["total_cost"].quantize(Decimal("0.01"))),
                "balance_value": str(balance_value.quantize(Decimal("0.01"))),
                "average_balance_cost": str(average_balance_cost.quantize(Decimal("0.01"))),
            })

        return Response(result)


class InventoryAdjustmentViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = InventoryAdjustmentSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]
    required_roles = [Membership.Role.OWNER, Membership.Role.ADMIN]

    def get_queryset(self):
        return (
            InventoryAdjustment.objects
            .select_related("warehouse", "item")
            .prefetch_related("allocations")
            .all()
        )


class StockTransferViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = StockTransferSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]
    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
        Membership.Role.MEMBER,
    ]

    def get_queryset(self):
        return (
            StockTransfer.objects
            .select_related("source_warehouse", "destination_warehouse", "item")
            .prefetch_related("allocations")
            .all()
        )


class KardexExportView(KardexView):
    def get(self, request):
        response = super().get(request)

        if response.status_code != 200:
            return response

        rows = response.data

        wb = Workbook()
        ws = wb.active
        ws.title = "Kardex"

        headers = [
            "Date",
            "Type",
            "Reference",
            "Entry Qty",
            "Exit Qty",
            "Balance Qty",
            "Unit Cost",
            "Movement Cost",
            "Balance Value",
            "Avg Balance Cost",
        ]

        add_tenant_report_header(ws, request.tenant)

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([
                row["date"],
                row["type"],
                row["reference"],
                row["entry_quantity"],
                row["exit_quantity"],
                row["balance_quantity"],
                row["unit_cost"],
                row["total_cost"],
                row["balance_value"],
                row["average_balance_cost"],
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        http_response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"kardex_{request.query_params.get('warehouse')}_{request.query_params.get('item')}.xlsx"

        http_response["Content-Disposition"] = f'attachment; filename="{filename}"'

        log_audit_event(
            request=request,
            action="export",
            entity="KardexExport",
            status_code=200,
            meta={
                "format": "xlsx",
                "params": request.query_params.dict(),
            },
        )

        return http_response


class CurrentStockExportView(CurrentStockView):
    def get(self, request):
        response = super().get(request)

        if response.status_code != 200:
            return response

        rows = response.data

        wb = Workbook()
        ws = wb.active
        ws.title = "Current Stock"

        headers = [
            "Warehouse",
            "Item Code",
            "Item Name",
            "Quantity",
            "Average Cost",
            "Total Cost",
        ]

        add_tenant_report_header(ws, request.tenant)

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([
                row["warehouse_name"],
                row["item_code"],
                row["item_name"],
                row["quantity"],
                row["average_cost"],
                row["total_cost"],
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        http_response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        http_response["Content-Disposition"] = 'attachment; filename="current_stock.xlsx"'

        log_audit_event(
            request=request,
            action="export",
            entity="CurrentStockExport",
            status_code=200,
            meta={
                "format": "xlsx",
            },
        )

        return http_response


class InventoryValuationView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember]

    def get(self, request):
        layer_value = ExpressionWrapper(
            F("remaining_quantity") * F("unit_cost"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )

        rows = (
            StockLayer.objects
            .values(
                "warehouse_id",
                "warehouse__name",
            )
            .annotate(
                inventory_value=Sum(layer_value),
            )
        )

        warehouses = []
        total_inventory_value = Decimal("0")

        for row in rows:
            value = row["inventory_value"] or Decimal("0")

            warehouses.append({
                "warehouse_id": row["warehouse_id"],
                "warehouse_name": row["warehouse__name"],
                "inventory_value": str(
                    value.quantize(Decimal("0.01"))
                ),
            })

            total_inventory_value += value

        return Response({
            "total_inventory_value": str(
                total_inventory_value.quantize(Decimal("0.01"))
            ),
            "warehouses": warehouses,
        })


class InventoryValuationExportView(InventoryValuationView):
    def get(self, request):
        response = super().get(request)

        if response.status_code != 200:
            return response

        data = response.data

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Valuation"

        headers = [
            "Warehouse",
            "Inventory Value",
        ]

        add_tenant_report_header(ws, request.tenant)

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for warehouse in data["warehouses"]:
            ws.append([
                warehouse["warehouse_name"],
                warehouse["inventory_value"],
            ])

        ws.append([])
        ws.append([
            "TOTAL INVENTORY VALUE",
            data["total_inventory_value"],
        ])

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        http_response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        http_response["Content-Disposition"] = (
            'attachment; filename="inventory_valuation.xlsx"'
        )

        log_audit_event(
            request=request,
            action="export",
            entity="InventoryValuationExport",
            status_code=200,
            meta={
                "format": "xlsx",
            },
        )

        return http_response


class InventoryValuationPdfView(InventoryValuationView):
    def get(self, request):
        response = super().get(request)

        if response.status_code != 200:
            return response

        pdf = build_inventory_valuation_pdf(
            tenant=request.tenant,
            data=response.data,
            user=request.user,
        )

        http_response = HttpResponse(
            pdf,
            content_type="application/pdf",
        )
        http_response["Content-Disposition"] = (
            'attachment; filename="inventory_valuation.pdf"'
        )

        log_audit_event(
            request=request,
            action="export",
            entity="InventoryValuationPdf",
            status_code=200,
            meta={
                "format": "pdf",
            },
        )

        return http_response


class CurrentStockPdfView(CurrentStockView):
    def get(self, request):
        response = super().get(request)

        if response.status_code != 200:
            return response

        pdf = build_current_stock_pdf(
            tenant=request.tenant,
            rows=response.data,
            user=request.user,
        )

        http_response = HttpResponse(
            pdf,
            content_type="application/pdf",
        )
        http_response["Content-Disposition"] = (
            'attachment; filename="current_stock.pdf"'
        )

        log_audit_event(
            request=request,
            action="export",
            entity="CurrentStockPdf",
            status_code=200,
            meta={
                "format": "pdf",
            },
        )

        return http_response


class KardexPdfView(KardexView):
    def get(self, request):
        response = super().get(request)

        if response.status_code != 200:
            return response

        pdf = build_kardex_pdf(
            tenant=request.tenant,
            rows=response.data,
            user=request.user,
        )

        http_response = HttpResponse(
            pdf,
            content_type="application/pdf",
        )
        http_response["Content-Disposition"] = (
            'attachment; filename="kardex.pdf"'
        )

        log_audit_event(
            request=request,
            action="export",
            entity="KardexPdf",
            status_code=200,
            meta={
                "format": "pdf",
                "params": request.query_params.dict(),
            },
        )

        return http_response
    

class UFVRateViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = UFVRateSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    def get_queryset(self):
        return UFVRate.objects.all()


class UFVRevaluationPreviewView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    def get(self, request):
        closing_date = request.query_params.get("closing_date")

        if not closing_date:
            return Response(
                {"detail": "closing_date query param is required."},
                status=400,
            )

        closing_ufv = UFVRate.objects.filter(date=closing_date).first()

        if not closing_ufv:
            return Response(
                {"detail": "UFV rate for closing_date was not found."},
                status=400,
            )

        rows = []
        total_original_value = Decimal("0")
        total_updated_value = Decimal("0")
        total_revaluation = Decimal("0")

        layers = (
            StockLayer.objects
            .select_related("warehouse", "item")
            .filter(
                remaining_quantity__gt=0,
                ufv_value__isnull=False,
            )
            .order_by("warehouse__name", "item__code", "entry_date", "id")
        )

        for layer in layers:
            original_total = layer.remaining_quantity * layer.unit_cost

            updated_unit_cost = layer.unit_cost * (
                closing_ufv.value / layer.ufv_value
            )

            updated_total = layer.remaining_quantity * updated_unit_cost
            revaluation_amount = updated_total - original_total

            total_original_value += original_total
            total_updated_value += updated_total
            total_revaluation += revaluation_amount

            rows.append({
                "warehouse_id": layer.warehouse_id,
                "warehouse_name": layer.warehouse.name,
                "item_id": layer.item_id,
                "item_code": layer.item.code,
                "item_name": layer.item.name,
                "entry_date": layer.entry_date,
                "quantity": str(layer.remaining_quantity.quantize(Decimal("0.01"))),
                "original_unit_cost": str(layer.unit_cost.quantize(Decimal("0.01"))),
                "purchase_ufv": str(layer.ufv_value.quantize(Decimal("0.00001"))),
                "closing_ufv": str(closing_ufv.value.quantize(Decimal("0.00001"))),
                "updated_unit_cost": str(updated_unit_cost.quantize(Decimal("0.01"))),
                "original_total": str(original_total.quantize(Decimal("0.01"))),
                "updated_total": str(updated_total.quantize(Decimal("0.01"))),
                "revaluation_amount": str(revaluation_amount.quantize(Decimal("0.01"))),
            })

        return Response({
            "closing_date": closing_date,
            "closing_ufv": str(closing_ufv.value.quantize(Decimal("0.00001"))),
            "total_original_value": str(total_original_value.quantize(Decimal("0.01"))),
            "total_updated_value": str(total_updated_value.quantize(Decimal("0.01"))),
            "total_revaluation": str(total_revaluation.quantize(Decimal("0.01"))),
            "rows": rows,
        })


class UFVRevaluationApplyView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    @transaction.atomic
    def post(self, request):
        closing_date = request.data.get("closing_date")
        notes = request.data.get("notes", "")

        if not closing_date:
            return Response(
                {"detail": "closing_date is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        closing_ufv = UFVRate.objects.filter(date=closing_date).first()

        if not closing_ufv:
            return Response(
                {"detail": "UFV rate for closing_date was not found."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        layers = (
            StockLayer.objects
            .select_related("warehouse", "item")
            .filter(
                remaining_quantity__gt=0,
                ufv_value__isnull=False,
            )
            .order_by("warehouse__name", "item__code", "entry_date", "id")
        )

        if not layers.exists():
            return Response(
                {"detail": "No stock layers with UFV value found."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        total_original_value = Decimal("0")
        total_updated_value = Decimal("0")
        total_revaluation = Decimal("0")
        calculated_lines = []

        for layer in layers:
            original_total = layer.remaining_quantity * layer.unit_cost
            updated_unit_cost = layer.unit_cost * (
                closing_ufv.value / layer.ufv_value
            )
            updated_total = layer.remaining_quantity * updated_unit_cost
            revaluation_amount = updated_total - original_total

            total_original_value += original_total
            total_updated_value += updated_total
            total_revaluation += revaluation_amount

            calculated_lines.append({
                "layer": layer,
                "quantity": layer.remaining_quantity,
                "original_unit_cost": layer.unit_cost,
                "updated_unit_cost": updated_unit_cost,
                "purchase_ufv": layer.ufv_value,
                "closing_ufv": closing_ufv.value,
                "original_total": original_total,
                "updated_total": updated_total,
                "revaluation_amount": revaluation_amount,
            })

        try:
            run = UFVRevaluationRun.objects.create(
                tenant=request.tenant,
                closing_date=closing_date,
                closing_ufv=closing_ufv.value,
                total_original_value=total_original_value,
                total_updated_value=total_updated_value,
                total_revaluation=total_revaluation,
                notes=notes,
                applied_by=request.user,
            )
        except IntegrityError:
            return Response(
                {
                    "detail": "UFV revaluation has already been applied for this closing date."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        for line in calculated_lines:
            layer = line["layer"]

            UFVRevaluationRunLine.objects.create(
                tenant=request.tenant,
                run=run,
                stock_layer=layer,
                warehouse=layer.warehouse,
                item=layer.item,
                quantity=line["quantity"],
                original_unit_cost=line["original_unit_cost"],
                updated_unit_cost=line["updated_unit_cost"],
                purchase_ufv=line["purchase_ufv"],
                closing_ufv=line["closing_ufv"],
                original_total=line["original_total"],
                updated_total=line["updated_total"],
                revaluation_amount=line["revaluation_amount"],
            )

        serializer = UFVRevaluationRunSerializer(run)

        log_audit_event(
            request=request,
            action="apply",
            entity="UFVRevaluationRun",
            entity_id=run.id,
            status_code=201,
            meta={
                "closing_date": closing_date,
                "total_revaluation": str(total_revaluation.quantize(Decimal("0.01"))),
            },
        )

        return Response(serializer.data, status=status.HTTP_201_CREATED)


class UFVRevaluationRunViewSet(TenantRequiredMixin, ReadOnlyModelViewSet):
    serializer_class = UFVRevaluationRunSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    def get_queryset(self):
        return (
            UFVRevaluationRun.objects
            .select_related("applied_by")
            .prefetch_related("lines")
            .all()
        )


class UFVRevaluationRunExportView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    def get(self, request, pk):
        run = (
            UFVRevaluationRun.objects
            .select_related("applied_by")
            .prefetch_related(
                "lines__warehouse",
                "lines__item",
            )
            .filter(pk=pk)
            .first()
        )

        if not run:
            return Response({"detail": "UFV revaluation run not found."}, status=404)

        wb = Workbook()
        ws = wb.active
        ws.title = "UFV Revaluation"

        add_tenant_report_header(ws, request.tenant)

        ws.append(["UFV Revaluation Run"])
        ws.append(["Closing Date", run.closing_date])
        ws.append(["Closing UFV", str(run.closing_ufv)])
        ws.append(["Applied By", run.applied_by.get_username() if run.applied_by else "-"])
        ws.append(["Created At", run.created_at.strftime("%Y-%m-%d %H:%M")])
        ws.append([])

        headers = [
            "Warehouse",
            "Item Code",
            "Item Name",
            "Quantity",
            "Purchase UFV",
            "Closing UFV",
            "Original Unit Cost",
            "Updated Unit Cost",
            "Original Total",
            "Updated Total",
            "Revaluation Amount",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for line in run.lines.all():
            ws.append([
                line.warehouse.name,
                line.item.code,
                line.item.name,
                str(line.quantity),
                str(line.purchase_ufv),
                str(line.closing_ufv),
                str(line.original_unit_cost),
                str(line.updated_unit_cost),
                str(line.original_total),
                str(line.updated_total),
                str(line.revaluation_amount),
            ])

        ws.append([])
        ws.append(["TOTAL ORIGINAL VALUE", str(run.total_original_value)])
        ws.append(["TOTAL UPDATED VALUE", str(run.total_updated_value)])
        ws.append(["TOTAL REVALUATION", str(run.total_revaluation)])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        log_audit_event(
            request=request,
            action="export",
            entity="UFVRevaluationRunExport",
            entity_id=run.id,
            status_code=200,
            meta={
                "format": "xlsx",
                "run_id": run.id,
                "closing_date": str(run.closing_date),
            },
        )

        http_response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        http_response["Content-Disposition"] = (
            f'attachment; filename="ufv_revaluation_run_{run.id}.xlsx"'
        )

        return http_response


class UFVRevaluationRunPdfView(TenantRequiredMixin, APIView):
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    def get(self, request, pk):
        run = (
            UFVRevaluationRun.objects
            .select_related("applied_by")
            .prefetch_related("lines__warehouse", "lines__item")
            .filter(pk=pk)
            .first()
        )

        if not run:
            return Response({"detail": "UFV revaluation run not found."}, status=404)

        pdf = build_ufv_revaluation_run_pdf(
            tenant=request.tenant,
            run=run,
            user=request.user,
        )

        log_audit_event(
            request=request,
            action="export",
            entity="UFVRevaluationRunPdf",
            entity_id=run.id,
            status_code=200,
            meta={
                "format": "pdf",
                "run_id": run.id,
                "closing_date": str(run.closing_date),
            },
        )

        http_response = HttpResponse(pdf, content_type="application/pdf")
        http_response["Content-Disposition"] = (
            f'attachment; filename="ufv_revaluation_run_{run.id}.pdf"'
        )

        return http_response


class StockEntryDocumentViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    serializer_class = StockEntryDocumentSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
        Membership.Role.MEMBER,
    ]

    def get_queryset(self):
        queryset = (
            StockEntryDocument.objects
            .prefetch_related("lines__warehouse", "lines__item")
            .all()
        )

        status_filter = self.request.query_params.get("status")
        supplier = self.request.query_params.get("supplier")
        document_number = self.request.query_params.get("document_number")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if supplier:
            queryset = queryset.filter(supplier_name__icontains=supplier)

        if document_number:
            queryset = queryset.filter(document_number__icontains=document_number)

        if date_from:
            queryset = queryset.filter(entry_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(entry_date__lte=date_to)

        return queryset.order_by("-entry_date", "-id")

    @transaction.atomic
    @action(detail=True, methods=["post"])
    def confirm(self, request, pk=None):
        document = self.get_object()

        if document.status != StockEntryDocument.Status.DRAFT:
            return Response(
                {"detail": "Only draft documents can be confirmed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lines = list(document.lines.select_related("warehouse", "item").all())

        if not lines:
            return Response(
                {"detail": "Document must have at least one line."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        total_amount = Decimal("0")

        for line in lines:
            StockLayer.objects.create(
                tenant=request.tenant,
                stock_entry=None,
                stock_entry_line=line,
                warehouse=line.warehouse,
                item=line.item,
                original_quantity=line.quantity,
                remaining_quantity=line.quantity,
                unit_cost=line.unit_cost,
                entry_date=document.entry_date,
                ufv_value=line.ufv_value,
            )

            total_amount += line.total_cost

        document.total_amount = total_amount
        document.status = StockEntryDocument.Status.CONFIRMED
        document.save(update_fields=["total_amount", "status"])

        log_audit_event(
            request=request,
            action="confirm",
            entity="StockEntryDocument",
            entity_id=document.id,
            status_code=200,
            meta={
                "document_number": document.document_number,
                "total_amount": str(total_amount),
            },
        )

        serializer = self.get_serializer(document)
        return Response(serializer.data)
    
    @transaction.atomic
    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        document = self.get_object()
        reason = request.data.get("reason", "")

        if document.status != StockEntryDocument.Status.CONFIRMED:
            return Response(
                {"detail": "Only confirmed documents can be cancelled."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lines = document.lines.select_related("warehouse", "item").prefetch_related(
            "stock_layer"
        )

        for line in lines:
            layer = getattr(line, "stock_layer", None)

            if not layer:
                return Response(
                    {"detail": f"Stock layer not found for line {line.id}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if layer.remaining_quantity != layer.original_quantity:
                return Response(
                    {
                        "detail": (
                            f"Cannot cancel entry document. "
                            f"Stock for item {line.item.code} has already been consumed."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        for line in lines:
            line.stock_layer.delete()

        document.status = StockEntryDocument.Status.CANCELLED
        document.cancelled_at = timezone.now()
        document.cancellation_reason = reason
        document.save(
            update_fields=[
                "status",
                "cancelled_at",
                "cancellation_reason",
            ]
        )

        log_audit_event(
            request=request,
            action="cancel",
            entity="StockEntryDocument",
            entity_id=document.id,
            status_code=200,
            meta={
                "document_number": document.document_number,
                "reason": reason,
            },
        )

        serializer = self.get_serializer(document)
        return Response(serializer.data)
    
    @action(detail=True, methods=["get"], url_path="download-pdf")
    def download_pdf(self, request, pk=None):
        document = self.get_object()

        if not document.document_pdf:
            return Response(
                {"detail": "Document PDF not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return FileResponse(
            document.document_pdf.open("rb"),
            content_type="application/pdf",
            as_attachment=False,
            filename=document.document_pdf.name.split("/")[-1],
        )
    
    @action(detail=True, methods=["get"])
    def pdf(self, request, pk=None):
        document = self.get_object()

        pdf = build_stock_entry_document_pdf(document)

        return FileResponse(
            pdf,
            content_type="application/pdf",
            filename=(
                f"stock-entry-"
                f"{document.document_number}.pdf"
            ),
        )


class StockExitDocumentViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    serializer_class = StockExitDocumentSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
        Membership.Role.MEMBER,
    ]

    def get_queryset(self):
        queryset = (
            StockExitDocument.objects
            .prefetch_related(
                "lines__warehouse",
                "lines__item",
                "lines__allocations",
            )
            .all()
        )

        status_filter = self.request.query_params.get("status")
        requester = self.request.query_params.get("requester")
        requesting_unit = self.request.query_params.get("requesting_unit")
        document_number = self.request.query_params.get("document_number")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if requester:
            queryset = queryset.filter(requester_name__icontains=requester)

        if requesting_unit:
            queryset = queryset.filter(requesting_unit__icontains=requesting_unit)

        if document_number:
            queryset = queryset.filter(document_number__icontains=document_number)

        if date_from:
            queryset = queryset.filter(exit_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(exit_date__lte=date_to)

        return queryset.order_by("-exit_date", "-id")

    @transaction.atomic
    @action(detail=True, methods=["post"])
    def confirm(self, request, pk=None):
        document = self.get_object()

        if document.status != StockExitDocument.Status.DRAFT:
            return Response(
                {"detail": "Only draft documents can be confirmed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lines = list(document.lines.select_related("warehouse", "item").all())

        if not lines:
            return Response(
                {"detail": "Document must have at least one line."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for line in lines:
            available = (
                StockLayer.objects
                .filter(
                    warehouse=line.warehouse,
                    item=line.item,
                    remaining_quantity__gt=0,
                )
                .aggregate(total=Sum("remaining_quantity"))
                .get("total")
                or Decimal("0")
            )

            if line.quantity > available:
                return Response(
                    {
                        "detail": (
                            f"Not enough stock for {line.item.code}. "
                            f"Available: {available}"
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        total_amount = Decimal("0")

        for line in lines:
            quantity_to_consume = line.quantity
            line_total_cost = Decimal("0")

            layers = (
                StockLayer.objects
                .select_for_update()
                .filter(
                    warehouse=line.warehouse,
                    item=line.item,
                    remaining_quantity__gt=0,
                )
                .order_by("entry_date", "id")
            )

            for layer in layers:
                if quantity_to_consume <= 0:
                    break

                consumed_quantity = min(
                    quantity_to_consume,
                    layer.remaining_quantity,
                )

                StockExitLineAllocation.objects.create(
                    tenant=request.tenant,
                    stock_exit_line=line,
                    stock_layer=layer,
                    quantity=consumed_quantity,
                    unit_cost=layer.unit_cost,
                )

                line_total_cost += consumed_quantity * layer.unit_cost

                layer.remaining_quantity -= consumed_quantity
                layer.save(update_fields=["remaining_quantity"])

                quantity_to_consume -= consumed_quantity

            line.total_cost = line_total_cost
            line.save(update_fields=["total_cost"])

            total_amount += line_total_cost

        document.total_amount = total_amount
        document.status = StockExitDocument.Status.CONFIRMED
        document.save(update_fields=["total_amount", "status"])

        log_audit_event(
            request=request,
            action="confirm",
            entity="StockExitDocument",
            entity_id=document.id,
            status_code=200,
            meta={
                "document_number": document.document_number,
                "total_amount": str(total_amount),
            },
        )

        serializer = self.get_serializer(document)
        return Response(serializer.data)
    
    @transaction.atomic
    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        document = self.get_object()
        reason = request.data.get("reason", "")

        if document.status != StockExitDocument.Status.CONFIRMED:
            return Response(
                {"detail": "Only confirmed documents can be cancelled."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lines = (
            document.lines
            .prefetch_related("allocations__stock_layer")
            .all()
        )

        for line in lines:
            for allocation in line.allocations.all():
                layer = allocation.stock_layer
                layer.remaining_quantity += allocation.quantity
                layer.save(update_fields=["remaining_quantity"])

        document.status = StockExitDocument.Status.CANCELLED
        document.cancelled_at = timezone.now()
        document.cancellation_reason = reason
        document.save(
            update_fields=[
                "status",
                "cancelled_at",
                "cancellation_reason",
            ]
        )

        log_audit_event(
            request=request,
            action="cancel",
            entity="StockExitDocument",
            entity_id=document.id,
            status_code=200,
            meta={
                "document_number": document.document_number,
                "reason": reason,
            },
        )

        serializer = self.get_serializer(document)
        return Response(serializer.data)
    
    @action(detail=True, methods=["get"], url_path="download-pdf")
    def download_pdf(self, request, pk=None):
        document = self.get_object()

        if not document.document_pdf:
            return Response(
                {"detail": "Document PDF not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return FileResponse(
            document.document_pdf.open("rb"),
            content_type="application/pdf",
            as_attachment=False,
            filename=document.document_pdf.name.split("/")[-1],
        )
    
    @action(detail=True, methods=["get"])
    def pdf(self, request, pk=None):
        document = self.get_object()

        pdf = build_stock_exit_document_pdf(document)

        return FileResponse(
            pdf,
            content_type="application/pdf",
            filename=f"stock-exit-{document.document_number}.pdf",
        )


class DocumentTypeViewSet(AuditCrudMixin, TenantRequiredMixin, ModelViewSet):
    serializer_class = DocumentTypeSerializer
    permission_classes = [IsAuthenticated, IsTenantMember, HasTenantRole]

    required_roles = [
        Membership.Role.OWNER,
        Membership.Role.ADMIN,
    ]

    def get_queryset(self):
        queryset = DocumentType.objects.all()

        movement_type = self.request.query_params.get("movement_type")
        is_active = self.request.query_params.get("is_active")

        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        return queryset
    
    @action(detail=False, methods=["post"], url_path="seed-defaults")
    def seed_defaults(self, request):
        defaults = [
            {
                "code": "FAC",
                "name": "Factura",
                "movement_type": DocumentType.MovementType.ENTRY,
                "requires_supplier": True,
                "requires_supplier_tax_id": True,
                "requires_requester": False,
                "requires_requesting_unit": False,
                "requires_pdf": True,
                "is_active": True,
            },
            {
                "code": "DON",
                "name": "Donación",
                "movement_type": DocumentType.MovementType.ENTRY,
                "requires_supplier": True,
                "requires_supplier_tax_id": False,
                "requires_requester": False,
                "requires_requesting_unit": False,
                "requires_pdf": False,
                "is_active": True,
            },
            {
                "code": "SAL",
                "name": "Vale de Salida",
                "movement_type": DocumentType.MovementType.EXIT,
                "requires_supplier": False,
                "requires_supplier_tax_id": False,
                "requires_requester": True,
                "requires_requesting_unit": True,
                "requires_pdf": False,
                "is_active": True,
            },
            {
                "code": "ACT",
                "name": "Acta",
                "movement_type": DocumentType.MovementType.BOTH,
                "requires_supplier": False,
                "requires_supplier_tax_id": False,
                "requires_requester": False,
                "requires_requesting_unit": False,
                "requires_pdf": True,
                "is_active": True,
            },
        ]

        created = 0

        for item in defaults:
            _, was_created = DocumentType.objects.get_or_create(
                tenant=request.tenant,
                code=item["code"],
                defaults=item,
            )

            if was_created:
                created += 1

        log_audit_event(
            request=request,
            action="seed",
            entity="DocumentType",
            status_code=200,
            meta={
                "created": created,
            },
        )

        return Response({
            "detail": "Default document types seeded.",
            "created": created,
        })
